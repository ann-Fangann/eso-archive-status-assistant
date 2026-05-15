from fastapi import FastAPI, UploadFile, HTTPException, File, Body, Form
from fastapi.responses import FileResponse
import uvicorn
from pydantic import BaseModel
import pandas as pd
import xlsxwriter
import os
import logging
from typing import List, Dict, Any
import json
import io
import re
from datetime import datetime, date, timedelta
from pathlib import Path
from uuid import uuid4
from sqlalchemy import create_engine, MetaData, Table, Column, inspect, text
from sqlalchemy.engine import URL
from sqlalchemy.types import Integer, String, Text, Float
from dotenv import load_dotenv
from openpyxl import load_workbook
import sys
import traceback

# 设置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# 定义允许的工作表名和字段
ALLOWED_SHEET_NAME = "工程车近14天数据"
REQUIRED_COLUMNS = [
    "vin", "统计日期", "日行里程", "累计里程", "日第一帧信号时间", 
    "日最后一帧信号时间", "日最后省份", "日最后城市", "日最后区域", 
    "日行程数", "日在线时长(小时)", "key_id"
]

app = FastAPI()

OUTPUT_DIR = Path(__file__).resolve().parent / "outputs"
OUTPUT_DIR.mkdir(exist_ok=True)

EMPTY_MARKERS = {"", "NA", "N/A", "NANA", "NONE", "NULL", "NAN", "NAT", "<NA>", "-", "--"}

ESO_MAIN_COLUMNS = [
    '操作类型', '层级', '数据类型', '产品', '车型年', '零件号', '短SVPPS', 'FFC',
    'FFC中文描述', '状态', '工程采购级别', '左右件', '功能组', '工程师代码',
    '工程师', '工厂编号', '节点组', '本色件标识', 'TG2_Plan_Date', 'TG2_Actual_Date',
    '2D_Drawing_Actual_Date', '2D_Drawing_No.', 'ESO_Plan_Date', 'ESO材料送检计划',
    'ESO备注_类型', '是否需要ESO_物流提供_', 'ESO_Actual_Date', '备注', '首次申请项目', '首次应用项目'
]

LAST_ANALYSIS: Dict[str, Any] = {
    "eso": None,
    "drawing": None,
}

@app.post("/upload-excel/")
async def upload_excel(file: UploadFile):
    # 读取Excel文件
    try:
        # 使用pandas读取Excel文件
        df = pd.read_excel(file.file)
        
        # 检查工作表名称
        if df.shape[0] == 0:
            raise HTTPException(status_code=400, detail="Excel文件为空")
            
        # 获取工作表名称（这里假设只有一个工作表）
        sheet_name = df.columns.name if df.columns.name else "Sheet1"
        if sheet_name != ALLOWED_SHEET_NAME:
            raise HTTPException(status_code=400, detail=f"工作表名称必须为'{ALLOWED_SHEET_NAME}'，当前为'{sheet_name}'")
            
        # 检查字段是否完整
        missing_columns = [col for col in REQUIRED_COLUMNS if col not in df.columns]
        if missing_columns:
            raise HTTPException(status_code=400, detail=f"缺少必要字段：{', '.join(missing_columns)}")
            
        # 添加字段说明
        field_descriptions = {
            "vin": "车辆唯一标识符",
            "统计日期": "数据统计的日期",
            "日行里程": "当日行驶的里程数",
            "累计里程": "累计行驶的总里程",
            "日第一帧信号时间": "当日第一次收到信号的时间",
            "日最后一帧信号时间": "当日最后一次收到信号的时间",
            "日最后省份": "当日最后所在省份",
            "日最后城市": "当日最后所在城市",
            "日最后区域": "当日最后所在区域",
            "日行程数": "当日行程次数",
            "日在线时长(小时)": "当日在线时长（单位：小时）",
            "key_id": "唯一标识ID"
        }
        
        # 返回结果
        return {
            "success": True,
            "message": "文件上传成功",
            "field_descriptions": field_descriptions,
            "data_preview": df.head().to_dict(),
            "total_rows": len(df)
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# 数据库配置。默认值保留公司内网配置，本地开发可在 .env 中覆盖。
load_dotenv()

DB_HOST = os.getenv("DB_HOST", "10.182.37.12")
DB_PORT = int(os.getenv("DB_PORT", "3306"))
DB_NAME = os.getenv("DB_NAME", "ESO")
DB_USER = os.getenv("DB_USER", "root")
DB_PASSWORD = os.getenv("DB_PASSWORD", "123456")
DB_CONNECT_TIMEOUT = int(os.getenv("DB_CONNECT_TIMEOUT", "3"))

# 使用utf8mb4字符集以支持中文
DATABASE_URL = URL.create(
    "mysql+pymysql",
    username=DB_USER,
    password=DB_PASSWORD,
    host=DB_HOST,
    port=DB_PORT,
    database=DB_NAME,
    query={"charset": "utf8mb4"},
)

logger.info(f"数据库配置: {DB_USER}@{DB_HOST}:{DB_PORT}/{DB_NAME}")

# 创建引擎时启用连接池预检查和支持utf8mb4
engine = create_engine(
    DATABASE_URL,
    echo=False,
    pool_pre_ping=True,
    pool_recycle=3600,
    connect_args={"connect_timeout": DB_CONNECT_TIMEOUT},
)

# 安全：处理表名和字段名，但保留中文字符
def sanitize_name(name: str, is_column=False) -> str:
    """
    清理名称以确保数据库兼容性
    对于表名：保留中文字符，但清理可能导致SQL问题的特殊字符
    对于列名：保留中文字符，但清理可能导致SQL问题的特殊字符
    """
    if not isinstance(name, str):
        name = str(name)
    
    # 对于表名和列名，都保留中文、英文字母、数字，移除可能导致问题的特殊字符
    name = re.sub(r'[^\w\u4e00-\u9fff]', '_', name)
    # 确保不以数字开头
    if name and name[0].isdigit():
        name = "t_" + name
    
    # 限制长度
    return name[:64] or "unnamed"

# 推断 SQL 类型 - 简化版本，全部使用字符串类型
def infer_sql_type(series):
    # 所有列都使用VARCHAR(255)类型以避免类型转换问题
    return String(255)

def create_table_with_chunked_columns(table_name, columns, type_mapping):
    """
    当列数过多时，将表按列分组创建多个子表
    """
    MAX_COLS_PER_TABLE = 50  # 每个表最多50列
    
    if len(columns) <= MAX_COLS_PER_TABLE:
        # 列数不多，直接创建一个表
        metadata = MetaData()
        table = Table(
            table_name, metadata, *columns,
            mysql_charset='utf8mb4',
            extend_existing=True,
            mysql_row_format='DYNAMIC',
            mysql_engine='InnoDB'
        )
        metadata.create_all(engine)
        return [table_name]  # 返回表名列表
    else:
        # 列数过多，需要拆分成多个表
        table_names = []
        chunk_num = 0
        
        # 主表包含前50列，包括关键标识列
        main_columns = columns[:MAX_COLS_PER_TABLE-1]  # 保留一列用于ID
        # 确保第一张表包含关键列（如果存在）
        essential_cols = [col for col in columns if '零件号' in col.name or 'ID' in col.name or 'id' in col.name]
        if essential_cols and main_columns[0] != essential_cols[0]:
            # 如果关键列不在前面，则调整顺序
            main_columns = [essential_cols[0]] + columns[1:MAX_COLS_PER_TABLE-1]
        
        chunk_table_name = f"{table_name}_part{chunk_num}"
        table_names.append(chunk_table_name)
        
        metadata = MetaData()
        main_table = Table(
            chunk_table_name, metadata, *main_columns,
            mysql_charset='utf8mb4',
            extend_existing=True,
            mysql_row_format='DYNAMIC',
            mysql_engine='InnoDB'
        )
        metadata.create_all(engine)
        chunk_num += 1
        
        # 处理其余的列
        remaining_columns = columns[MAX_COLS_PER_TABLE-1:]
        for i in range(0, len(remaining_columns), MAX_COLS_PER_TABLE-1):  # 保留一列用于ID关联
            chunk = remaining_columns[i:i+MAX_COLS_PER_TABLE-1]
            chunk_table_name = f"{table_name}_part{chunk_num}"
            table_names.append(chunk_table_name)
            
            # 为了关联，每个分表都需要一个ID列
            id_col = Column('common_id', Integer, primary_key=True)
            chunk_with_id = [id_col] + chunk
            
            metadata = MetaData()
            part_table = Table(
                chunk_table_name, metadata, *chunk_with_id,
                mysql_charset='utf8mb4',
                extend_existing=True,
                mysql_row_format='DYNAMIC',
                mysql_engine='InnoDB'
            )
            metadata.create_all(engine)
            chunk_num += 1
            
        return table_names

def convert_to_int(value):
    """安全地将值转换为整数"""
    if pd.isna(value) or value == '' or str(value).strip() == '':
        return None
    try:
        # 先转换为浮点数，再转换为整数
        float_val = float(value)
        if float_val.is_integer():
            return int(float_val)
        else:
            # 如果不是整数，取整
            return int(float_val)
    except (ValueError, TypeError):
        # 如果不能转换为整数，返回None
        return None

def convert_to_float(value):
    """安全地将值转换为浮点数"""
    if pd.isna(value) or value == '' or str(value).strip() == '':
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        # 如果不能转换为浮点数，返回None
        return None

def process_date_value(value):
    """
    处理日期值，只保留年月日部分
    """
    if pd.isna(value) or value == '' or value == 'None' or value == 'nan':
        return value

    # 尝试解析为pandas的Timestamp或datetime对象
    try:
        if isinstance(value, (pd.Timestamp, datetime, date)):
            # 只保留年月日部分
            if isinstance(value, (pd.Timestamp, datetime)):
                return value.strftime('%Y-%m-%d')
            else:  # date对象
                return value.strftime('%Y-%m-%d')
    except:
        pass

    # 尝试解析字符串格式的日期（可能包含时间部分）
    try:
        value_str = str(value)
        # 匹配常见的日期格式：YYYY-MM-DD HH:MM:SS 或 YYYY/MM/DD HH:MM:SS 等
        if ' ' in value_str and ('-' in value_str[:10] or '/' in value_str[:10]):
            # 分割日期和时间部分
            date_part = value_str.split(' ')[0]
            # 验证是否是有效的日期格式
            if '-' in date_part and len(date_part) == 10:
                return date_part
            elif '/' in date_part:
                # 将 YYYY/MM/DD 转换为 YYYY-MM-DD
                parts = date_part.split('/')
                if len(parts) == 3:
                    return f"{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
    except:
        pass

    # 如果不是日期格式，返回原值
    return value

def preprocess_data(df, type_mapping):
    """预处理数据，将所有数据转换为字符串并处理空值，日期只保留年月日"""
    for col in df.columns:
        # 先处理日期类型 - 如果列包含日期数据，只保留年月日部分
        if df[col].dtype == 'object':  # 对象类型可能包含日期字符串
            df[col] = df[col].apply(lambda x: process_date_value(x))

        # 将所有数据转换为字符串，处理各种空值情况
        df[col] = df[col].apply(lambda x: '' if pd.isna(x) or x == 'nan' or x == '<NA>' or x == 'None' else str(x))
        
        # 截断过长的字符串以避免行大小问题
        if isinstance(type_mapping.get(col), String) and hasattr(type_mapping.get(col), 'length'):
            max_length = type_mapping[col].length
            df[col] = df[col].apply(lambda x: x[:max_length] if len(x) > max_length else x)
        else:
            # 对于其他类型，限制为较小的长度
            df[col] = df[col].apply(lambda x: x[:255] if len(x) > 255 else x)
    
    return df

def is_effective_value(value) -> bool:
    """判断业务字段是否有有效值，排除空值和 NA/NANA 等无需 ESO 标记。"""
    if value is None:
        return False
    try:
        if pd.isna(value):
            return False
    except (TypeError, ValueError):
        pass
    value_str = str(value).strip()
    return value_str.upper() not in EMPTY_MARKERS

def normalize_part_no(value) -> str:
    """零件号匹配用标准化，避免 Excel 数字/空格导致匹配失败。"""
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    value_str = str(value).strip()
    if value_str.endswith(".0"):
        value_str = value_str[:-2]
    return value_str

def parse_date_like_value(value):
    """将常见 Excel/字符串日期转成 date/datetime，无法识别时返回原值。"""
    if value is None or not is_effective_value(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if isinstance(value, (datetime, date)):
        return value

    if isinstance(value, (int, float)) and 1 <= float(value) <= 60000:
        try:
            return pd.to_datetime(value, unit="D", origin="1899-12-30").to_pydatetime()
        except Exception:
            pass

    value_str = str(value).strip()
    value_str = value_str.split(" ")[0].replace("/", "-")
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return datetime.strptime(value_str, fmt)
        except ValueError:
            continue
    parsed = pd.to_datetime(value, errors="coerce")
    if not pd.isna(parsed):
        return parsed.to_pydatetime()
    return value

def archive_date_sort_key(value):
    """归档日期去重比较键，无法识别的日期排在最小。"""
    parsed = parse_date_like_value(value)
    if isinstance(parsed, pd.Timestamp):
        return parsed.date()
    if isinstance(parsed, datetime):
        return parsed.date()
    if isinstance(parsed, date):
        return parsed
    return date.min

def resolve_target_date(target_date: str | None) -> date:
    """统计日期默认昨天；用户传入时按 YYYY-MM-DD 校验。"""
    if not target_date or not target_date.strip():
        return date.today() - timedelta(days=1)
    try:
        return datetime.strptime(target_date.strip(), "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="统计日期格式错误，应为 YYYY-MM-DD")

def quote_identifier(name: str) -> str:
    """为 MySQL 标识符加反引号，避免中文列名或特殊列名造成 SQL 歧义。"""
    return f"`{str(name).replace('`', '``')}`"

def build_duplicate_part_stats(conn, table_name: str, table_columns: List[str]):
    """统计重复零件号，先暴露风险，不在业务规则未确认前擅自去重。"""
    if '零件号' not in table_columns:
        return {
            "duplicate_part_count": 0,
            "duplicate_row_count": 0,
            "samples": [],
        }

    table_ident = quote_identifier(table_name)
    part_ident = quote_identifier('零件号')
    duplicate_summary_sql = f"""
    SELECT
        COUNT(*) AS duplicate_part_count,
        COALESCE(SUM(row_count), 0) AS duplicate_row_count
    FROM (
        SELECT {part_ident}, COUNT(*) AS row_count
        FROM {table_ident}
        WHERE {part_ident} IS NOT NULL AND TRIM({part_ident}) != ''
        GROUP BY {part_ident}
        HAVING COUNT(*) > 1
    ) duplicate_parts;
    """
    duplicate_samples_sql = f"""
    SELECT {part_ident} AS 零件号, COUNT(*) AS count
    FROM {table_ident}
    WHERE {part_ident} IS NOT NULL AND TRIM({part_ident}) != ''
    GROUP BY {part_ident}
    HAVING COUNT(*) > 1
    ORDER BY count DESC
    LIMIT 20;
    """

    summary = conn.execute(text(duplicate_summary_sql)).mappings().first() or {}
    samples = [dict(row) for row in conn.execute(text(duplicate_samples_sql)).mappings().all()]
    return {
        "duplicate_part_count": int(summary.get("duplicate_part_count") or 0),
        "duplicate_row_count": int(summary.get("duplicate_row_count") or 0),
        "samples": samples,
    }

def build_unfinished_select_result(conn, resolved_target_date: date, matched_completed_count=None):
    """基于当前 sheet 表，按统计日期生成未完成清单和汇总。"""
    try:
        main_table_columns = conn.execute(text("SHOW COLUMNS FROM `sheet`")).fetchall()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"请先上传并处理 Excel 文件，再生成未完成清单: {str(e)}")

    main_columns = [col[0] for col in main_table_columns]
    required_main_cols = ['零件号', 'ESO_Plan_Date', 'ESO_Actual_Date']
    missing_main_cols = [col for col in required_main_cols if col not in main_columns]
    if missing_main_cols:
        raise HTTPException(status_code=400, detail=f"主表(sheet)缺少必要列: {missing_main_cols}")

    plan_col = quote_identifier('ESO_Plan_Date')
    actual_col = quote_identifier('ESO_Actual_Date')
    operation_col = quote_identifier('操作类型')
    has_operation_col = '操作类型' in main_columns
    active_row_condition = (
        f"({operation_col} IS NULL OR UPPER(TRIM({operation_col})) != 'D')"
        if has_operation_col else "1 = 1"
    )
    delete_row_condition = (
        f"({operation_col} IS NOT NULL AND UPPER(TRIM({operation_col})) = 'D')"
        if has_operation_col else "0 = 1"
    )

    plan_valid_condition = """
    {plan_col} IS NOT NULL
    AND TRIM({plan_col}) != ''
    AND UPPER(TRIM({plan_col})) NOT IN ('NA', 'N/A', 'NANA', 'NONE', 'NULL', 'NAN')
    """.format(plan_col=plan_col)
    plan_date_expr = f"""
    STR_TO_DATE(
        REPLACE(SUBSTRING_INDEX(TRIM({plan_col}), ' ', 1), '/', '-'),
        '%Y-%m-%d'
    )
    """
    plan_due_condition = f"""
    {plan_valid_condition}
    AND {plan_date_expr} IS NOT NULL
    AND {plan_date_expr} <= :target_date
    """
    actual_filled_condition = f"""
    {actual_col} IS NOT NULL
    AND TRIM({actual_col}) != ''
    """
    actual_empty_condition = f"NOT ({actual_filled_condition})"
    active_plan_due_unfinished_condition = f"""
    {active_row_condition}
    AND {plan_due_condition}
    AND {actual_empty_condition}
    """

    summary_sql = f"""
    SELECT
        SUM(CASE WHEN {active_row_condition} AND {plan_valid_condition} THEN 1 ELSE 0 END) AS planned_count,
        SUM(CASE WHEN {active_row_condition} AND {plan_valid_condition} AND {actual_filled_condition} THEN 1 ELSE 0 END) AS completed_count,
        SUM(CASE WHEN {active_plan_due_unfinished_condition} THEN 1 ELSE 0 END) AS unfinished_count,
        SUM(CASE WHEN {active_row_condition} AND {plan_due_condition} THEN 1 ELSE 0 END) AS due_planned_count,
        SUM(CASE WHEN {active_row_condition} AND {plan_due_condition} AND {actual_filled_condition} THEN 1 ELSE 0 END) AS due_completed_count,
        SUM(CASE WHEN {delete_row_condition} THEN 1 ELSE 0 END) AS delete_row_count,
        SUM(CASE WHEN {delete_row_condition} AND {actual_filled_condition} THEN 1 ELSE 0 END) AS delete_completed_count,
        SUM(CASE WHEN {delete_row_condition} AND {plan_due_condition} AND {actual_empty_condition} THEN 1 ELSE 0 END) AS delete_unfinished_excluded_count
    FROM sheet;
    """

    group_col = '部门' if '部门' in main_columns else '功能组' if '功能组' in main_columns else None
    if group_col:
        group_ident = quote_identifier(group_col)
        group_expr = f"COALESCE(NULLIF(TRIM({group_ident}), ''), '未知{group_col}')"
    else:
        group_expr = "'未知功能组'"

    group_sql = f"""
    SELECT
        {group_expr} AS 功能组,
        COUNT(*) AS count
    FROM sheet
    WHERE {active_plan_due_unfinished_condition}
    GROUP BY {group_expr}
    ORDER BY count DESC;
    """

    select_columns = "*, '审批中' AS `ESO状态`" if 'ESO状态' not in main_columns else "*"
    select_sql = f"""
    SELECT {select_columns}
    FROM sheet
    WHERE {active_plan_due_unfinished_condition};
    """

    query_params = {"target_date": resolved_target_date.isoformat()}
    summary_result = conn.execute(text(summary_sql), query_params).mappings().first() or {}
    planned_count = int(summary_result.get("planned_count") or 0)
    completed_count = int(summary_result.get("completed_count") or 0)
    unfinished_count = int(summary_result.get("unfinished_count") or 0)
    due_planned_count = int(summary_result.get("due_planned_count") or 0)
    due_completed_count = int(summary_result.get("due_completed_count") or 0)
    delete_row_count = int(summary_result.get("delete_row_count") or 0)
    delete_completed_count = int(summary_result.get("delete_completed_count") or 0)
    delete_unfinished_excluded_count = int(summary_result.get("delete_unfinished_excluded_count") or 0)

    group_result = conn.execute(text(group_sql), query_params)
    group_rows = [dict(row) for row in group_result.mappings().all()]

    result = conn.execute(text(select_sql), query_params)
    columns = result.keys()
    rows = [dict(zip(columns, row)) for row in result.fetchall()]
    logger.info(f"未完成清单查询成功，统计日期 {resolved_target_date.isoformat()}，返回 {len(rows)} 条记录")
    duplicate_stats = {
        "sheet": build_duplicate_part_stats(conn, "sheet", main_columns),
    }
    try:
        join_table_columns = conn.execute(text("SHOW COLUMNS FROM `sheet1`")).fetchall()
        duplicate_stats["sheet1"] = build_duplicate_part_stats(conn, "sheet1", [col[0] for col in join_table_columns])
    except Exception:
        duplicate_stats["sheet1"] = {
            "duplicate_part_count": 0,
            "duplicate_row_count": 0,
            "samples": [],
        }

    return {
        "row_count": len(rows),
        "total_empty_count": unfinished_count,
        "data": rows,
        "summary": {
            "planned_count": planned_count,
            "completed_count": completed_count,
            "unfinished_count": unfinished_count,
            "due_planned_count": due_planned_count,
            "due_completed_count": due_completed_count,
            "delete_row_count": delete_row_count,
            "delete_completed_count": delete_completed_count,
            "delete_unfinished_excluded_count": delete_unfinished_excluded_count,
            "matched_completed_count": matched_completed_count,
            "unfinished_by_difference": max(due_planned_count - due_completed_count, 0),
            "target_date": resolved_target_date.isoformat(),
            "formula": "ESO Plan Date <= 统计日期，ESO Actual Date 为空，且操作类型不是 D = 未完成数量",
            "group_field": group_col or "功能组",
        },
        "group_stats": group_rows,
        "duplicate_stats": duplicate_stats,
    }

def find_column_by_sanitized_header(ws, header_row: int, expected_name: str):
    for cell in ws[header_row]:
        if sanitize_name(cell.value, is_column=True) == expected_name:
            return cell.column
    return None

def build_archive_date_map(file_bytes: bytes):
    """
    从 TDSP 已归档清单中读取零件号 -> 归档日期。
    返回原始 openpyxl 单元格值和格式，便于写回时尽量保留日期显示格式。
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    part_col = find_column_by_sanitized_header(ws, 1, "零件号")
    archive_col = find_column_by_sanitized_header(ws, 1, "归档日期")

    if not part_col or not archive_col:
        raise ValueError("ESO已归档清单缺少必要列：零件号或归档日期")

    archive_map = {}
    for row in range(2, ws.max_row + 1):
        part_no = normalize_part_no(ws.cell(row=row, column=part_col).value)
        archive_cell = ws.cell(row=row, column=archive_col)
        if part_no and is_effective_value(archive_cell.value):
            archive_info = {
                "value": archive_cell.value,
                "number_format": archive_cell.number_format,
            }
            existing_info = archive_map.get(part_no)
            if not existing_info or archive_date_sort_key(archive_info["value"]) >= archive_date_sort_key(existing_info["value"]):
                archive_map[part_no] = archive_info
    return archive_map

def create_modified_club_workbook(file_bytes: bytes, archive_map: Dict, original_filename: str):
    """
    在原始零件俱乐部 workbook 上回填 ESO Actual Date。
    只写实际日期列，不触碰 ESO Plan Date，因此计划日期格式会保留原样。
    """
    wb = load_workbook(io.BytesIO(file_bytes))
    ws = wb[wb.sheetnames[0]]
    header_row = 2
    part_col = find_column_by_sanitized_header(ws, header_row, "零件号")
    actual_col = find_column_by_sanitized_header(ws, header_row, "ESO_Actual_Date")
    operation_col = find_column_by_sanitized_header(ws, header_row, "操作类型")

    if not part_col or not actual_col:
        raise ValueError("零件俱乐部清单缺少必要列：零件号或 ESO Actual Date")

    filled_count = 0
    delete_skipped_count = 0
    for row in range(header_row + 1, ws.max_row + 1):
        part_no = normalize_part_no(ws.cell(row=row, column=part_col).value)
        archive_info = archive_map.get(part_no)
        if not archive_info:
            continue

        operation_value = ""
        if operation_col:
            operation_value = str(ws.cell(row=row, column=operation_col).value or "").strip().upper()
        if operation_value == "D":
            delete_skipped_count += 1
            continue

        actual_cell = ws.cell(row=row, column=actual_col)
        actual_cell.value = parse_date_like_value(archive_info["value"])
        actual_cell.number_format = "yyyy/mm/dd"
        filled_count += 1

    safe_stem = re.sub(r"[^\w\u4e00-\u9fff.-]+", "_", Path(original_filename).stem)
    output_name = f"{safe_stem}-已回填ESO实际归档日期-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid4().hex[:8]}.xlsx"
    output_path = OUTPUT_DIR / output_name
    wb.save(output_path)

    return {
        "filename": output_name,
        "download_url": f"/download/{output_name}",
        "filled_count": filled_count,
        "delete_skipped_count": delete_skipped_count,
    }

def to_jsonable_value(value):
    """把 pandas/openpyxl 值转成 FastAPI 可安全返回的基础类型。"""
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(value, pd.Timestamp):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            pass
    return value

def format_date_cell_value(value):
    parsed = parse_date_like_value(value)
    if isinstance(parsed, (datetime, date, pd.Timestamp)):
        return to_jsonable_value(parsed)
    return "" if not is_effective_value(parsed) else str(parsed).strip()

def dataframe_to_records(df: pd.DataFrame) -> List[Dict[str, Any]]:
    records = []
    for _, row in df.iterrows():
        records.append({col: to_jsonable_value(row.get(col)) for col in df.columns})
    return records

def duplicate_part_stats_from_dataframe(df: pd.DataFrame):
    if df is None or df.empty or '零件号' not in df.columns:
        return {
            "duplicate_part_count": 0,
            "duplicate_row_count": 0,
            "samples": [],
        }

    part_numbers = df['零件号'].apply(normalize_part_no)
    counts = part_numbers[part_numbers != ""].value_counts()
    duplicates = counts[counts > 1]
    return {
        "duplicate_part_count": int(len(duplicates)),
        "duplicate_row_count": int(duplicates.sum()) if len(duplicates) else 0,
        "samples": [
            {"零件号": part_no, "count": int(count)}
            for part_no, count in duplicates.head(20).items()
        ],
    }

def read_eso_main_dataframe(file_bytes: bytes):
    excel_file = io.BytesIO(file_bytes)
    sheet_names = pd.ExcelFile(excel_file).sheet_names
    if not sheet_names:
        raise HTTPException(status_code=400, detail="零件俱乐部文件没有工作表")

    sheet_name = sheet_names[0]
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, header=1)
    if df.empty:
        raise HTTPException(status_code=400, detail="零件俱乐部第一个工作表为空")

    original_columns = list(df.columns)
    eso_column_names = [col for col in original_columns if isinstance(col, str) and 'ESO' in col.upper()]
    df.columns = deduplicate_columns([sanitize_name(col, is_column=True) for col in df.columns])

    available_columns = [col for col in ESO_MAIN_COLUMNS if col in df.columns]
    eso_columns_in_required = {'ESO_Plan_Date', 'ESO_Actual_Date', 'ESO备注_类型', '是否需要ESO_物流提供_'}
    for original_col in eso_column_names:
        sanitized_name = sanitize_name(original_col, is_column=True)
        if sanitized_name in eso_columns_in_required:
            continue
        if sanitized_name in df.columns and sanitized_name not in available_columns:
            available_columns.append(sanitized_name)
        elif sanitized_name not in df.columns:
            for col in df.columns:
                if (col.startswith(sanitized_name + '_') or sanitized_name in col) and col not in available_columns:
                    available_columns.append(col)
                    break

    for first_project_col in ('首次申请项目', '首次应用项目'):
        if first_project_col in df.columns and first_project_col not in available_columns:
            available_columns.append(first_project_col)
    if not available_columns:
        available_columns = df.columns.tolist()

    return df[available_columns].copy(), sheet_names

def read_archive_dataframe(file_bytes: bytes):
    excel_file = io.BytesIO(file_bytes)
    sheet_names = pd.ExcelFile(excel_file).sheet_names
    if not sheet_names:
        raise HTTPException(status_code=400, detail="ESO零件清单文件没有工作表")

    sheet_name = sheet_names[0]
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name)
    if df.empty:
        return df, sheet_names

    df.columns = deduplicate_columns([sanitize_name(col, is_column=True) for col in df.columns])
    return df, sheet_names

def apply_archive_to_eso_dataframe(df: pd.DataFrame, archive_map: Dict):
    if '零件号' not in df.columns or 'ESO_Actual_Date' not in df.columns:
        return 0, 0

    filled_count = 0
    delete_skipped_count = 0
    for idx, row in df.iterrows():
        part_no = normalize_part_no(row.get('零件号'))
        archive_info = archive_map.get(part_no)
        if not archive_info:
            continue

        operation_value = str(row.get('操作类型') or "").strip().upper() if '操作类型' in df.columns else ""
        if operation_value == "D":
            delete_skipped_count += 1
            continue

        df.at[idx, 'ESO_Actual_Date'] = format_date_cell_value(archive_info["value"])
        filled_count += 1

    return filled_count, delete_skipped_count

def build_eso_select_result_from_dataframe(
    df: pd.DataFrame,
    resolved_target_date: date,
    matched_completed_count=None,
    archive_df: pd.DataFrame | None = None,
):
    required_cols = ['零件号', 'ESO_Plan_Date', 'ESO_Actual_Date']
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        raise HTTPException(status_code=400, detail=f"零件俱乐部缺少必要列: {missing_cols}")

    rows = []
    planned_count = 0
    completed_count = 0
    unfinished_count = 0
    due_planned_count = 0
    due_completed_count = 0
    delete_row_count = 0
    delete_completed_count = 0
    delete_unfinished_excluded_count = 0
    group_counter: Dict[str, int] = {}
    group_col = '部门' if '部门' in df.columns else '功能组' if '功能组' in df.columns else None

    for _, row in df.iterrows():
        operation_value = str(row.get('操作类型') or "").strip().upper() if '操作类型' in df.columns else ""
        is_delete = operation_value == "D"
        plan_date = parse_date_to_date(row.get('ESO_Plan_Date'))
        has_valid_plan = plan_date is not None and is_effective_value(row.get('ESO_Plan_Date'))
        actual_filled = is_effective_value(row.get('ESO_Actual_Date'))
        plan_due = has_valid_plan and plan_date <= resolved_target_date

        if is_delete:
            delete_row_count += 1
            if actual_filled:
                delete_completed_count += 1
            if plan_due and not actual_filled:
                delete_unfinished_excluded_count += 1
            continue

        if has_valid_plan:
            planned_count += 1
            if actual_filled:
                completed_count += 1
        if plan_due:
            due_planned_count += 1
            if actual_filled:
                due_completed_count += 1

        if plan_due and not actual_filled:
            unfinished_count += 1
            output_row = {col: to_jsonable_value(row.get(col)) for col in df.columns}
            output_row["ESO状态"] = "审批中"
            rows.append(output_row)
            group_name = str(output_row.get(group_col) or f"未知{group_col or '功能组'}")
            group_counter[group_name] = group_counter.get(group_name, 0) + 1

    group_rows = [
        {"功能组": key, "count": value}
        for key, value in sorted(group_counter.items(), key=lambda item: item[1], reverse=True)
    ]
    duplicate_stats = {
        "sheet": duplicate_part_stats_from_dataframe(df),
        "sheet1": duplicate_part_stats_from_dataframe(archive_df) if archive_df is not None else {
            "duplicate_part_count": 0,
            "duplicate_row_count": 0,
            "samples": [],
        },
    }

    return {
        "row_count": len(rows),
        "total_empty_count": unfinished_count,
        "data": rows,
        "summary": {
            "planned_count": planned_count,
            "completed_count": completed_count,
            "unfinished_count": unfinished_count,
            "due_planned_count": due_planned_count,
            "due_completed_count": due_completed_count,
            "delete_row_count": delete_row_count,
            "delete_completed_count": delete_completed_count,
            "delete_unfinished_excluded_count": delete_unfinished_excluded_count,
            "matched_completed_count": matched_completed_count,
            "unfinished_by_difference": max(due_planned_count - due_completed_count, 0),
            "target_date": resolved_target_date.isoformat(),
            "formula": "ESO Plan Date <= 统计日期，ESO Actual Date 为空，且操作类型不是 D = 未完成数量",
            "group_field": group_col or "功能组",
        },
        "group_stats": group_rows,
        "duplicate_stats": duplicate_stats,
    }

def ensure_standard_tables(conn):
    conn.execute(text("""
    CREATE TABLE IF NOT EXISTS analysis_batches (
        batch_id VARCHAR(64) PRIMARY KEY,
        scene VARCHAR(32) NOT NULL,
        target_date VARCHAR(20),
        file1_name VARCHAR(255),
        file2_name VARCHAR(255),
        summary_json LONGTEXT,
        created_at DATETIME NOT NULL
    ) CHARACTER SET utf8mb4;
    """))
    conn.execute(text("""
    CREATE TABLE IF NOT EXISTS analysis_unfinished_items (
        id BIGINT AUTO_INCREMENT PRIMARY KEY,
        batch_id VARCHAR(64) NOT NULL,
        scene VARCHAR(32) NOT NULL,
        row_index INT NOT NULL,
        part_no VARCHAR(255),
        group_name VARCHAR(255),
        engineer VARCHAR(255),
        item_type VARCHAR(255),
        status VARCHAR(255),
        payload_json LONGTEXT,
        created_at DATETIME NOT NULL,
        INDEX idx_batch_scene (batch_id, scene),
        INDEX idx_scene_part (scene, part_no)
    ) CHARACTER SET utf8mb4;
    """))
    conn.execute(text("""
    CREATE TABLE IF NOT EXISTS analysis_group_stats (
        id BIGINT AUTO_INCREMENT PRIMARY KEY,
        batch_id VARCHAR(64) NOT NULL,
        scene VARCHAR(32) NOT NULL,
        group_name VARCHAR(255),
        count_value INT NOT NULL,
        created_at DATETIME NOT NULL,
        INDEX idx_batch_scene (batch_id, scene)
    ) CHARACTER SET utf8mb4;
    """))

def persist_standard_result(scene: str, file1_name: str, file2_name: str | None, select_result: Dict[str, Any]):
    """MySQL 只做标准结果存储；连接失败不影响 Excel 处理主流程。"""
    batch_id = uuid4().hex
    now = datetime.now()
    try:
        with engine.begin() as conn:
            ensure_standard_tables(conn)
            conn.execute(text("""
            INSERT INTO analysis_batches
                (batch_id, scene, target_date, file1_name, file2_name, summary_json, created_at)
            VALUES
                (:batch_id, :scene, :target_date, :file1_name, :file2_name, :summary_json, :created_at)
            """), {
                "batch_id": batch_id,
                "scene": scene,
                "target_date": select_result.get("summary", {}).get("target_date"),
                "file1_name": file1_name,
                "file2_name": file2_name,
                "summary_json": json.dumps(select_result.get("summary", {}), ensure_ascii=False, default=str),
                "created_at": now,
            })

            item_rows = []
            for idx, row in enumerate(select_result.get("data", []), start=1):
                item_rows.append({
                    "batch_id": batch_id,
                    "scene": scene,
                    "row_index": idx,
                    "part_no": normalize_part_no(row.get("零件号")),
                    "group_name": row.get("功能组") or row.get("部门") or "",
                    "engineer": row.get("工程师") or "",
                    "item_type": row.get("未完成类型") or "ESO未完成",
                    "status": row.get("ESO状态") or row.get("图纸状态") or "审批中",
                    "payload_json": json.dumps(row, ensure_ascii=False, default=str),
                    "created_at": now,
                })
            if item_rows:
                conn.execute(text("""
                INSERT INTO analysis_unfinished_items
                    (batch_id, scene, row_index, part_no, group_name, engineer, item_type, status, payload_json, created_at)
                VALUES
                    (:batch_id, :scene, :row_index, :part_no, :group_name, :engineer, :item_type, :status, :payload_json, :created_at)
                """), item_rows)

            stat_rows = [
                {
                    "batch_id": batch_id,
                    "scene": scene,
                    "group_name": item.get("功能组") or item.get("部门") or item.get("未完成类型") or "",
                    "count_value": int(item.get("count") or item.get("延期未发布数量") or 0),
                    "created_at": now,
                }
                for item in select_result.get("group_stats", [])
            ]
            if stat_rows:
                conn.execute(text("""
                INSERT INTO analysis_group_stats
                    (batch_id, scene, group_name, count_value, created_at)
                VALUES
                    (:batch_id, :scene, :group_name, :count_value, :created_at)
                """), stat_rows)

        return {"saved": True, "batch_id": batch_id, "message": "结果已保存到 MySQL，智能问答可查询历史批次"}
    except Exception as e:
        logger.warning("标准结果保存到 MySQL 失败，主流程继续: %s", str(e))
        return {"saved": False, "batch_id": None, "error": str(e), "message": "MySQL不可用，结果仅保存在当前服务内存中"}

def process_eso_files(file1_bytes: bytes, file2_bytes: bytes, file1_name: str, file2_name: str, resolved_target_date: date):
    main_df, sheet_names1 = read_eso_main_dataframe(file1_bytes)
    archive_df, sheet_names2 = read_archive_dataframe(file2_bytes)
    archive_map = build_archive_date_map(file2_bytes)
    modified_workbook = create_modified_club_workbook(file1_bytes, archive_map, file1_name)
    updated_row_count, delete_skipped_count = apply_archive_to_eso_dataframe(main_df, archive_map)
    modified_workbook["filled_count"] = updated_row_count
    modified_workbook["delete_skipped_count"] = delete_skipped_count

    select_result = build_eso_select_result_from_dataframe(
        main_df,
        resolved_target_date,
        matched_completed_count=updated_row_count,
        archive_df=archive_df,
    )
    persistence = persist_standard_result("eso", file1_name, file2_name, select_result)
    results = {
        "sheet": f"成功处理 {len(main_df)} 行零件俱乐部数据",
        "sheet1": f"成功处理 {len(archive_df)} 行ESO零件清单数据",
    }
    LAST_ANALYSIS["eso"] = {
        "scene": "eso",
        "df": main_df.copy(),
        "archive_df": archive_df.copy(),
        "file1": file1_name,
        "file2": file2_name,
        "sheet_names": sheet_names1 + sheet_names2,
        "results": results,
        "modified_workbook": modified_workbook,
        "persistence": persistence,
        "select_result": select_result,
    }
    return sheet_names1, sheet_names2, results, modified_workbook, select_result, persistence

def split_dataframe_for_tables(df, table_names, max_cols_per_table=49):  # 保留1列给ID
    """
    将DataFrame按列拆分以适应多个表
    """
    if len(table_names) == 1:
        # 只有一个表，不需要拆分
        return [(table_names[0], df)]
    
    dataframes_parts = []
    
    # 第一部分 - 主表
    main_df = df.iloc[:, :max_cols_per_table]
    dataframes_parts.append((table_names[0], main_df))
    
    # 其余部分 - 按照分表策略拆分
    remaining_df = df.iloc[:, max_cols_per_table:]
    
    part_num = 1
    for i in range(0, remaining_df.shape[1], max_cols_per_table):
        if part_num >= len(table_names):
            break
            
        chunk_df = remaining_df.iloc[:, i:i+max_cols_per_table].copy()
        # 添加一个公共ID列用于关联
        chunk_df['common_id'] = chunk_df.index + 1
        dataframes_parts.append((table_names[part_num], chunk_df))
        part_num += 1
    
    return dataframes_parts

def append_dataframe_to_table(table_name: str, df: pd.DataFrame):
    """使用 SQLAlchemy 原生批量插入，避免 pandas.to_sql 在不同服务器版本下不兼容。"""
    if df.empty:
        return

    metadata = MetaData()
    table = Table(table_name, metadata, autoload_with=engine)
    records = df.to_dict(orient="records")
    with engine.begin() as conn:
        conn.execute(table.insert(), records)

def deduplicate_columns(columns: List[str]) -> List[str]:
    seen = set()
    new_cols = []
    for col in columns:
        suffix = 1
        orig = col
        while col in seen:
            col = f"{orig}_{suffix}"
            suffix += 1
        seen.add(col)
        new_cols.append(col)
    return new_cols

def read_excel_with_header_candidates(file_bytes: bytes, header_candidates=(1, 0)) -> pd.DataFrame:
    """图纸模板还未固化，优先按零件俱乐部第二行表头读取，失败再退回第一行。"""
    last_df = None
    for header in header_candidates:
        excel_file = io.BytesIO(file_bytes)
        df = pd.read_excel(excel_file, sheet_name=0, header=header)
        if df.empty:
            last_df = df
            continue
        df.columns = deduplicate_columns([sanitize_name(col, is_column=True) for col in df.columns])
        if '零件号' in df.columns:
            return df
        last_df = df
    return last_df if last_df is not None else pd.DataFrame()

def find_first_existing_column(columns: List[str], aliases: List[str]):
    alias_set = {sanitize_name(alias, is_column=True) for alias in aliases}
    for col in columns:
        if col in alias_set:
            return col

    lowered_aliases = [sanitize_name(alias, is_column=True).lower() for alias in aliases]
    for col in columns:
        lowered_col = col.lower()
        if any(alias and alias in lowered_col for alias in lowered_aliases):
            return col
    return None

def parse_date_to_date(value):
    parsed = parse_date_like_value(value)
    if isinstance(parsed, datetime):
        return parsed.date()
    if isinstance(parsed, date):
        return parsed
    return None

def add_one_month(value: date) -> date:
    return (pd.Timestamp(value) + pd.DateOffset(months=1)).date()

DRAWING_COLUMN_ALIASES = {
    "part_no": ["零件号", "Part_No", "Part_Number", "Part Number"],
    "operation": ["操作类型"],
    "group": ["部门", "功能组", "专业组"],
    "engineer": ["工程师", "设计工程师", "负责人"],
    "model_plan": ["数模计划日期", "数模_Plan_Date", "数模PlanDate", "TG2_Plan_Date"],
    "model_actual": ["数模实际日期", "数模发布日期", "数模_Actual_Date", "数模ActualDate", "TG2_Actual_Date"],
    "drawing_plan": ["图纸计划日期", "图纸_Plan_Date", "图纸PlanDate", "2D_Drawing_Plan_Date"],
    "drawing_actual": ["图纸实际日期", "图纸发布日期", "图纸发布实际日期", "图纸_Actual_Date", "2D_Drawing_Actual_Date"],
    "drawing_no": ["图纸号", "2D_Drawing_No.", "2D_Drawing_No_", "t_2D_Drawing_No_"],
}

def detect_drawing_columns(columns: List[str]):
    return {
        key: find_first_existing_column(columns, aliases)
        for key, aliases in DRAWING_COLUMN_ALIASES.items()
    }

def build_drawing_actual_map(file_bytes: bytes):
    if not file_bytes:
        return {}

    df = read_excel_with_header_candidates(file_bytes, header_candidates=(0, 1))
    if df.empty:
        return {}

    columns = df.columns.tolist()
    detected = detect_drawing_columns(columns)
    part_col = detected.get("part_no")
    actual_col = detected.get("drawing_actual") or find_first_existing_column(columns, ["归档日期", "发布日期", "发布时间"])
    if not part_col or not actual_col:
        return {}

    actual_map = {}
    for _, row in df.iterrows():
        part_no = normalize_part_no(row.get(part_col))
        actual_value = row.get(actual_col)
        if part_no and is_effective_value(actual_value):
            existing_value = actual_map.get(part_no)
            if existing_value is None or archive_date_sort_key(actual_value) >= archive_date_sort_key(existing_value):
                actual_map[part_no] = actual_value
    return actual_map

def find_column_by_aliases(ws, header_row: int, aliases: List[str]):
    alias_set = {sanitize_name(alias, is_column=True) for alias in aliases}
    for cell in ws[header_row]:
        sanitized = sanitize_name(cell.value, is_column=True)
        if sanitized in alias_set:
            return cell.column

    lowered_aliases = [alias.lower() for alias in alias_set]
    for cell in ws[header_row]:
        sanitized = sanitize_name(cell.value, is_column=True).lower()
        if any(alias and alias in sanitized for alias in lowered_aliases):
            return cell.column
    return None

def detect_workbook_header_row(ws, candidates=(2, 1)):
    for header_row in candidates:
        part_col = find_column_by_aliases(ws, header_row, DRAWING_COLUMN_ALIASES["part_no"])
        if part_col:
            return header_row
    return candidates[0]

def create_modified_drawing_workbook(file_bytes: bytes, published_actual_map: Dict, original_filename: str):
    """
    在图纸/数模主清单上回填图纸实际日期。
    若原表没有图纸实际日期列，则在末尾新增“图纸实际日期”列。
    """
    wb = load_workbook(io.BytesIO(file_bytes))
    ws = wb[wb.sheetnames[0]]
    header_row = detect_workbook_header_row(ws)
    part_col = find_column_by_aliases(ws, header_row, DRAWING_COLUMN_ALIASES["part_no"])
    actual_col = find_column_by_aliases(ws, header_row, DRAWING_COLUMN_ALIASES["drawing_actual"])
    operation_col = find_column_by_aliases(ws, header_row, DRAWING_COLUMN_ALIASES["operation"])

    if not part_col:
        raise ValueError("图纸主清单缺少必要列：零件号")

    if not actual_col:
        actual_col = ws.max_column + 1
        ws.cell(row=header_row, column=actual_col).value = "图纸实际日期"

    filled_count = 0
    delete_skipped_count = 0
    for row in range(header_row + 1, ws.max_row + 1):
        part_no = normalize_part_no(ws.cell(row=row, column=part_col).value)
        if not part_no:
            continue

        operation_value = ""
        if operation_col:
            operation_value = str(ws.cell(row=row, column=operation_col).value or "").strip().upper()
        if operation_value == "D":
            delete_skipped_count += 1
            continue

        actual_value = published_actual_map.get(part_no)
        if not is_effective_value(actual_value):
            continue

        actual_cell = ws.cell(row=row, column=actual_col)
        if is_effective_value(actual_cell.value):
            continue
        actual_cell.value = parse_date_like_value(actual_value)
        actual_cell.number_format = "yyyy/mm/dd"
        filled_count += 1

    safe_stem = re.sub(r"[^\w\u4e00-\u9fff.-]+", "_", Path(original_filename).stem)
    output_name = f"{safe_stem}-已回填图纸实际日期-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid4().hex[:8]}.xlsx"
    output_path = OUTPUT_DIR / output_name
    wb.save(output_path)

    return {
        "filename": output_name,
        "download_url": f"/download/{output_name}",
        "filled_count": filled_count,
        "delete_skipped_count": delete_skipped_count,
    }

def build_drawing_unfinished_result(club_file_bytes: bytes, published_file_bytes: bytes | None, resolved_target_date: date):
    df = read_excel_with_header_candidates(club_file_bytes)
    if df.empty:
        raise HTTPException(status_code=400, detail="图纸零件俱乐部文件为空或无法读取")

    detected = detect_drawing_columns(df.columns.tolist())
    part_col = detected.get("part_no")
    model_plan_col = detected.get("model_plan")
    model_actual_col = detected.get("model_actual")
    drawing_actual_col = detected.get("drawing_actual")

    missing = []
    if not part_col:
        missing.append("零件号")
    if not model_plan_col:
        missing.append("数模计划日期/TG2 Plan Date")
    if not model_actual_col:
        missing.append("数模实际日期/TG2 Actual Date")
    if not drawing_actual_col and not published_file_bytes:
        missing.append("图纸实际发布日期/第二个图纸发布清单")
    if missing:
        raise HTTPException(
            status_code=400,
            detail={
                "message": f"图纸模板缺少必要字段: {', '.join(missing)}",
                "detected_columns": detected,
                "available_columns": df.columns.tolist(),
            },
        )

    type_mapping = {col: String(255) for col in df.columns}
    df = preprocess_data(df, type_mapping)
    published_actual_map = build_drawing_actual_map(published_file_bytes) if published_file_bytes else {}
    rows = []
    delete_excluded_count = 0
    model_due_count = 0
    drawing_due_count = 0
    model_completed_count = 0
    drawing_completed_count = 0

    operation_col = detected.get("operation")
    group_col = detected.get("group")
    engineer_col = detected.get("engineer")
    drawing_plan_col = detected.get("drawing_plan")
    drawing_no_col = detected.get("drawing_no")

    for _, row in df.iterrows():
        operation_value = str(row.get(operation_col, "") or "").strip().upper() if operation_col else ""
        if operation_value == "D":
            delete_excluded_count += 1
            continue

        part_no = normalize_part_no(row.get(part_col))
        if not part_no:
            continue

        model_plan_date = parse_date_to_date(row.get(model_plan_col))
        model_actual_date = parse_date_to_date(row.get(model_actual_col))
        drawing_plan_date = parse_date_to_date(row.get(drawing_plan_col)) if drawing_plan_col else None
        drawing_actual_value = row.get(drawing_actual_col) if drawing_actual_col else None
        if not is_effective_value(drawing_actual_value):
            drawing_actual_value = published_actual_map.get(part_no)
        drawing_actual_date = parse_date_to_date(drawing_actual_value)

        if model_actual_date:
            model_completed_count += 1
        if drawing_actual_date:
            drawing_completed_count += 1

        base_row = row.to_dict()
        base_row.update({
            "零件号": part_no,
            "功能组": row.get(group_col, "") if group_col else "",
            "工程师": row.get(engineer_col, "") if engineer_col else "",
            "图纸号": row.get(drawing_no_col, "") if drawing_no_col else "",
            "图纸实际日期": drawing_actual_date.isoformat() if drawing_actual_date else "",
        })

        if model_plan_date and model_plan_date <= resolved_target_date:
            model_due_count += 1

        if model_plan_date and model_plan_date <= resolved_target_date and not model_actual_date:
            rows.append({
                **base_row,
                "未完成类型": "数模未完成",
                "数模状态": "计划已到期未完成",
                "图纸状态": "待数模完成后一个月内发布",
                "图纸要求完成日期": "",
            })
            rows.append({
                **base_row,
                "未完成类型": "图纸未发布",
                "数模状态": "数模未完成",
                "图纸状态": "受数模未完成影响",
                "图纸要求完成日期": "待数模完成后一个月",
            })
            continue

        drawing_due_date = None
        if model_actual_date:
            drawing_due_date = add_one_month(model_actual_date)
        elif drawing_plan_date:
            drawing_due_date = drawing_plan_date

        if drawing_due_date and drawing_due_date <= resolved_target_date:
            drawing_due_count += 1

        if drawing_due_date and drawing_due_date <= resolved_target_date and not drawing_actual_date:
            rows.append({
                **base_row,
                "未完成类型": "图纸未发布",
                "数模状态": "数模已完成" if model_actual_date else "按图纸计划日期判断",
                "图纸状态": "到期未发布",
                "图纸要求完成日期": drawing_due_date.isoformat() if isinstance(drawing_due_date, date) else str(drawing_due_date),
            })

    group_counter: Dict[str, int] = {}
    type_counter: Dict[str, int] = {}
    for row in rows:
        group_name = str(row.get("功能组") or "未知功能组")
        type_name = str(row.get("未完成类型") or "未知类型")
        group_counter[group_name] = group_counter.get(group_name, 0) + 1
        type_counter[type_name] = type_counter.get(type_name, 0) + 1

    group_stats = [{"功能组": key, "count": value} for key, value in sorted(group_counter.items(), key=lambda item: item[1], reverse=True)]
    type_stats = [{"未完成类型": key, "count": value} for key, value in sorted(type_counter.items(), key=lambda item: item[1], reverse=True)]

    return {
        "row_count": len(rows),
        "data": rows,
        "group_stats": group_stats,
        "type_stats": type_stats,
        "summary": {
            "target_date": resolved_target_date.isoformat(),
            "unfinished_count": len(rows),
            "delete_excluded_count": delete_excluded_count,
            "model_due_count": model_due_count,
            "model_completed_count": model_completed_count,
            "drawing_due_count": drawing_due_count,
            "drawing_completed_count": drawing_completed_count,
            "formula": "数模计划日期 <= 统计日期且数模未完成，或数模完成后一个月内图纸未发布；操作类型 D 不进入未完成清单",
            "detected_columns": detected,
        },
    }

def process_drawing_files(
    club_file_bytes: bytes,
    published_file_bytes: bytes | None,
    club_filename: str,
    published_filename: str | None,
    resolved_target_date: date,
):
    published_actual_map = build_drawing_actual_map(published_file_bytes) if published_file_bytes else {}
    modified_workbook = None
    if published_actual_map:
        try:
            modified_workbook = create_modified_drawing_workbook(club_file_bytes, published_actual_map, club_filename)
        except Exception as workbook_error:
            logger.error(f"生成回填后的图纸主清单失败: {str(workbook_error)}", exc_info=True)
            modified_workbook = {
                "error": f"生成回填后的图纸主清单失败: {str(workbook_error)}"
            }
    else:
        modified_workbook = {
            "error": "未提供可识别的图纸发布/归档日期清单，未生成回填文件"
        }

    select_result = build_drawing_unfinished_result(club_file_bytes, published_file_bytes, resolved_target_date)
    if isinstance(modified_workbook, dict) and "filled_count" in modified_workbook:
        select_result["summary"]["matched_completed_count"] = modified_workbook.get("filled_count", 0)
        select_result["summary"]["delete_skipped_count"] = modified_workbook.get("delete_skipped_count", 0)

    persistence = persist_standard_result(
        "drawing",
        club_filename,
        published_filename,
        select_result,
    )
    LAST_ANALYSIS["drawing"] = {
        "scene": "drawing",
        "file1": club_filename,
        "file2": published_filename,
        "club_bytes": club_file_bytes,
        "published_bytes": published_file_bytes,
        "modified_workbook": modified_workbook,
        "select_result": select_result,
        "persistence": persistence,
    }
    return select_result, modified_workbook, persistence

@app.post("/upload-excel/", summary="上传Excel文件")
async def upload_excel(file: UploadFile = File(...)):
    """
    上传Excel文件并将其数据导入MySQL数据库
    
    - Excel工作表名称将作为数据库表名
    - 第一行将作为数据库字段名
    - 会删除已存在的同名表
    """
    logger.info(f"收到上传请求，文件名: {file.filename}")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        logger.warning(f"不支持的文件类型: {file.filename}")
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .xls 文件")

    try:
        logger.info(f"开始处理文件: {file.filename}")
        contents = await file.read()
        logger.info(f"文件大小: {len(contents)} bytes")
        excel_file = io.BytesIO(contents)

        # 读取所有 sheet
        sheet_names = pd.ExcelFile(excel_file).sheet_names
        logger.info(f"发现工作表: {sheet_names}")
        results = {}

        for sheet_name in sheet_names:
            logger.info(f"处理工作表: {sheet_name}")
            # 重置指针
            excel_file.seek(0)
            # 读取数据，第一行作为列名
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            logger.info(f"工作表 {sheet_name} 数据形状: {df.shape}")

            if df.empty:
                results[sheet_name] = "跳过（空表）"
                continue

            # 使用原始的第一行作为列名，但需要清理以支持中文
            df.columns = [sanitize_name(col, is_column=True) for col in df.columns]
            logger.info(f"清理后的列名: {df.columns.tolist()}")
            
            # 去重列名（避免重复）
            seen = set()
            new_cols = []
            for col in df.columns:
                suffix = 1
                orig = col
                while col in seen:
                    col = f"{orig}_{suffix}"
                    suffix += 1
                seen.add(col)
                new_cols.append(col)
            df.columns = new_cols

            # 表名使用工作表名称
            table_name = sanitize_name(sheet_name, is_column=False)
            logger.info(f"表名: {table_name}")
            
            # 获取数据库元数据
            metadata = MetaData()
            inspector = inspect(engine)

            # 删除已存在的同名表
            if inspector.has_table(table_name):
                logger.info(f"删除已存在的表: {table_name}")
                old_table = Table(table_name, metadata, autoload_with=engine)
                old_table.drop(engine)
                logger.info(f"成功删除表: {table_name}")

            # 推断列类型并构建表结构
            columns = []
            type_mapping = {}
            
            for col in df.columns:
                # 推断合适的SQL类型而不是全部使用TEXT
                sql_type = infer_sql_type(df[col])
                columns.append(Column(col, sql_type))
                # 构建 SQLAlchemy 类型映射（用于 to_sql）
                type_mapping[col] = sql_type

            # 创建新表 - 使用分表策略
            created_table_names = create_table_with_chunked_columns(table_name, columns, type_mapping)
            logger.info(f"成功创建表: {created_table_names}")

            # 数据预处理 - 确保所有数据都符合对应的SQL类型
            df = preprocess_data(df, type_mapping)

            # 插入数据
            logger.info(f"准备插入 {len(df)} 行数据到表 {table_name}")
            
            # 按表拆分数据
            table_dataframes = split_dataframe_for_tables(df, created_table_names)
            
            for t_name, t_df in table_dataframes:
                logger.info(f"准备插入 {len(t_df)} 行数据到表 {t_name}")
                # 获取当前表的类型映射
                current_type_mapping = {}
                for col_name in t_df.columns:
                    if col_name in type_mapping:
                        current_type_mapping[col_name] = type_mapping[col_name]
                
                # 数据预处理 - 确保所有数据都符合对应的SQL类型
                t_df = preprocess_data(t_df, current_type_mapping)
                
                # 分批插入数据，避免一次性插入大量数据
                batch_size = 1000
                total_rows = len(t_df)
                
                for start_idx in range(0, total_rows, batch_size):
                    end_idx = min(start_idx + batch_size, total_rows)
                    batch_df = t_df.iloc[start_idx:end_idx]
                    
                    logger.info(f"正在插入批次 {start_idx//batch_size + 1}/{(total_rows-1)//batch_size + 1} 到表 {t_name}")
                    
                    append_dataframe_to_table(t_name, batch_df)
                
                logger.info(f"成功插入 {len(t_df)} 行数据到表 {t_name}")

            results[sheet_name] = f"成功导入 {len(df)} 行数据到表 `{', '.join(created_table_names)}`"
            logger.info(f"成功插入数据到表: {created_table_names}")

        logger.info(f"文件处理完成: {file.filename}")
        return {
            "filename": file.filename,
            "sheets": len(sheet_names),
            "results": results
        }

    except Exception as e:
        logger.error(f"处理失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"处理失败: {str(e)}")

@app.post("/upload-two-excel/", summary="上传两个Excel文件并生成ESO未完成清单")
async def upload_two_excel(
    file1: UploadFile = File(...),
    file2: UploadFile = File(...),
    target_date: str = Form(None),
):
    """
    上传两个Excel文件并生成 ESO 未完成清单。

    处理主流程不依赖 MySQL；MySQL 可用时只用于保存标准结果，供智能问答跨重启查询。
    
    - 第一个文件从第二行开始读取列名，并按会议确认字段生成 ESO 清单
    - 第二个文件读取零件号与归档日期，用于回填 ESO Actual Date
    - 操作类型为 D 的行不回填、不进入未完成清单
    - 返回结构保留 sql_results 字段，用于兼容现有前端展示
    """
    logger.info(f"收到两个上传请求，文件1: {file1.filename}, 文件2: {file2.filename}")
    
    if not file1.filename.endswith(('.xlsx', '.xls')):
        logger.warning(f"不支持的文件类型: {file1.filename}")
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .xls 文件 (file1)")

    if not file2.filename.endswith(('.xlsx', '.xls')):
        logger.warning(f"不支持的文件类型: {file2.filename}")
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .xls 文件 (file2)")

    try:
        resolved_target_date = resolve_target_date(target_date)
        logger.info(f"未完成清单统计日期: {resolved_target_date.isoformat()}")

        contents1 = await file1.read()
        contents2 = await file2.read()
        sheet_names1, sheet_names2, results, modified_workbook, select_result, persistence = process_eso_files(
            contents1,
            contents2,
            file1.filename,
            file2.filename,
            resolved_target_date,
        )

        logger.info(
            "ESO文件处理完成，未完成 %s 条，MySQL保存状态: %s",
            select_result.get("row_count"),
            persistence.get("saved"),
        )
        return {
            "file1": file1.filename,
            "file2": file2.filename,
            "target_date": resolved_target_date.isoformat(),
            "sheets": len(sheet_names1) + len(sheet_names2),
            "results": results,
            "modified_workbook": modified_workbook,
            "persistence": persistence,
            "sql_results": {
                "update_result": f"成功按Python规则回填 {modified_workbook.get('filled_count', 0)} 条记录",
                "select_result": select_result
            }
        }

        # 处理第一个文件 - 强制导入为 'sheet' 表，但只导入指定列，从第二行开始读取列名
        logger.info(f"开始处理文件1: {file1.filename}")
        contents1 = await file1.read()
        logger.info(f"文件1大小: {len(contents1)} bytes")
        excel_file1 = io.BytesIO(contents1)

        # 读取所有 sheet
        sheet_names1 = pd.ExcelFile(excel_file1).sheet_names
        logger.info(f"文件1发现工作表: {sheet_names1}")
        results1 = {}

        # 定义第一个文件要保留的列
        required_columns_sheet1 = [
            '操作类型', '层级', '数据类型', '产品', '车型年', '零件号', '短SVPPS', 'FFC',
            'FFC中文描述', '状态', '工程采购级别', '左右件', '功能组', '工程师代码',
            '工程师', '工厂编号', '节点组', '本色件标识', 'TG2_Plan_Date', 'TG2_Actual_Date',
            '2D_Drawing_Actual_Date', '2D_Drawing_No.', 'ESO_Plan_Date', 'ESO材料送检计划',
            'ESO备注_类型', '是否需要ESO_物流提供_', 'ESO_Actual_Date', '备注', '首次申请项目'
        ]

        # 强制将第一个文件的第一个工作表导入为 'sheet' 表
        for idx, sheet_name in enumerate(sheet_names1):
            logger.info(f"开始处理工作表: {sheet_name}")
            # 重置指针
            excel_file1.seek(0)
            # 读取数据，从第二行开始读取列名（header=1）
            logger.info("开始读取Excel文件...")
            df = pd.read_excel(excel_file1, sheet_name=sheet_name, header=1)
            logger.info(f"工作表 {sheet_name} 数据形状: {df.shape}")

            if df.empty:
                results1[sheet_name] = "跳过（空表）"
                logger.info(f"工作表 {sheet_name} 为空，跳过处理")
                continue

            logger.info("开始清理列名...")

            # 在sanitize之前，先记录原始列名并检测包含"ESO"的列名
            original_columns = list(df.columns)
            eso_column_names = []
            for col in original_columns:
                if isinstance(col, str) and 'ESO' in col.upper():
                    eso_column_names.append(col)
                    logger.info(f"发现包含ESO的原始列: {col}")

            # 使用第二行作为列名，但需要清理以支持中文
            df.columns = [sanitize_name(col, is_column=True) for col in df.columns]
            logger.info(f"清理后的列名: {df.columns.tolist()}")

            # 去重列名（避免重复）
            seen = set()
            new_cols = []
            old_to_new_col_map = {}  # 记录列名映射
            for col in df.columns:
                suffix = 1
                orig = col
                old_to_new_col_map[col] = col  # 先记录原始映射
                while col in seen:
                    col = f"{orig}_{suffix}"
                    suffix += 1
                seen.add(col)
                old_to_new_col_map[orig] = col  # 更新映射
                new_cols.append(col)
            df.columns = new_cols
            logger.info(f"去重后的列名: {df.columns.tolist()}")

            logger.info("开始筛选指定列...")
            # 只保留指定的列
            available_columns = [col for col in required_columns_sheet1 if col in df.columns]

            # 添加所有包含"ESO"的列（先sanitize原始列名，再在去重后的列中查找）
            # 跳过已经在required_columns_sheet1中定义的ESO列
            eso_columns_in_required = ['ESO_Plan_Date', 'ESO_Actual_Date', 'ESO备注_类型', '是否需要ESO_物流提供_']

            for original_col in eso_column_names:
                sanitized_name = sanitize_name(original_col, is_column=True)
                # 如果这个列已经在required_columns_sheet1中定义，跳过
                if sanitized_name in eso_columns_in_required:
                    continue
                # 先尝试直接找sanitized后的名称
                if sanitized_name in df.columns and sanitized_name not in available_columns:
                    available_columns.append(sanitized_name)
                    logger.info(f"添加包含ESO的列: {sanitized_name} (原名: {original_col})")
                # 如果直接找不到，可能在去重时被加了后缀，尝试查找
                elif sanitized_name not in df.columns:
                    for col in df.columns:
                        if col.startswith(sanitized_name + '_') or sanitized_name in col:
                            if col not in available_columns:
                                available_columns.append(col)
                                logger.info(f"添加包含ESO的列: {col} (原名: {original_col})")
                            break

            # 添加首次申请项目列（如果存在且未包含）
            if '首次申请项目' in df.columns and '首次申请项目' not in available_columns:
                available_columns.append('首次申请项目')
                logger.info(f"添加首次申请项目列")

            if not available_columns:
                logger.warning(f"没有找到任何指定的列，将使用所有可用列")
                available_columns = df.columns.tolist()
            else:
                logger.info(f"将保留的列: {available_columns}")

            df = df[available_columns]
            logger.info(f"筛选后的数据形状: {df.shape}")

            # 强制将第一个文件的第一个工作表导入为 'sheet' 表
            table_name = "sheet" if idx == 0 else f"sheet_{idx}"
            logger.info(f"表名: {table_name}")
            
            # 获取数据库元数据
            metadata = MetaData()
            inspector = inspect(engine)

            # 删除已存在的同名表
            if inspector.has_table(table_name):
                logger.info(f"删除已存在的表: {table_name}")
                old_table = Table(table_name, metadata, autoload_with=engine)
                old_table.drop(engine)
                logger.info(f"成功删除表: {table_name}")

            logger.info("开始推断列类型并构建表结构...")
            # 推断列类型并构建表结构
            columns = []
            type_mapping = {}
            
            for col in df.columns:
                # 推断合适的SQL类型而不是全部使用TEXT
                sql_type = infer_sql_type(df[col])
                columns.append(Column(col, sql_type))
                # 构建 SQLAlchemy 类型映射（用于 to_sql）
                type_mapping[col] = sql_type

            # 创建新表
            table = Table(
                table_name, metadata, *columns,
                mysql_charset='utf8mb4',
                extend_existing=True,
                mysql_row_format='DYNAMIC',
                mysql_engine='InnoDB'
            )
            logger.info(f"创建表 {table_name}")
            metadata.create_all(engine)
            logger.info(f"成功创建表: {table_name}")

            logger.info("开始数据预处理...")
            # 数据预处理 - 确保所有数据都符合对应的SQL类型
            df = preprocess_data(df, type_mapping)
            logger.info("数据预处理完成")

            logger.info("开始插入数据...")
            # 插入数据
            logger.info(f"准备插入 {len(df)} 行数据到表 {table_name}")
            
            # 按表拆分数据 - 由于我们已经筛选了列，不再需要拆分
            logger.info(f"插入 {len(df)} 行数据到表 {table_name}")
            
            # 获取当前表的类型映射
            current_type_mapping = {}
            for col_name in df.columns:
                if col_name in type_mapping:
                    current_type_mapping[col_name] = type_mapping[col_name]
            
            # 分批插入数据，避免一次性插入大量数据
            batch_size = 1000
            total_rows = len(df)
            
            for start_idx in range(0, total_rows, batch_size):
                end_idx = min(start_idx + batch_size, total_rows)
                batch_df = df.iloc[start_idx:end_idx]
                
                logger.info(f"正在插入批次 {start_idx//batch_size + 1}/{(total_rows-1)//batch_size + 1} 到表 {table_name}")
                
                append_dataframe_to_table(table_name, batch_df)
            
            results1[sheet_name] = f"成功导入 {len(df)} 行数据到表 `{table_name}`"
            logger.info(f"成功插入数据到表: {table_name}")
            
            # 只处理第一个工作表，其他工作表跳过
            if idx == 0:
                break

        # 处理第二个文件 - 强制导入为 'sheet1' 表，保持原有逻辑
        logger.info(f"开始处理文件2: {file2.filename}")
        contents2 = await file2.read()
        logger.info(f"文件2大小: {len(contents2)} bytes")
        excel_file2 = io.BytesIO(contents2)
        modified_workbook = None

        try:
            archive_map = build_archive_date_map(contents2)
            modified_workbook = create_modified_club_workbook(contents1, archive_map, file1.filename)
            logger.info(
                "已生成回填后的零件俱乐部文件: %s，回填 %s 行，跳过 Delete 行 %s 行",
                modified_workbook["filename"],
                modified_workbook["filled_count"],
                modified_workbook["delete_skipped_count"],
            )
        except Exception as workbook_error:
            logger.error(f"生成回填后的零件俱乐部文件失败: {str(workbook_error)}", exc_info=True)
            modified_workbook = {
                "error": f"生成回填后的零件俱乐部文件失败: {str(workbook_error)}"
            }

        # 读取所有 sheet
        sheet_names2 = pd.ExcelFile(excel_file2).sheet_names
        logger.info(f"文件2发现工作表: {sheet_names2}")
        results2 = {}

        # 强制将第二个文件导入为 'sheet1' 表
        for idx, sheet_name in enumerate(sheet_names2):
            logger.info(f"开始处理工作表: {sheet_name}")
            # 重置指针
            excel_file2.seek(0)
            # 读取数据，第一行作为列名（保持原有逻辑）
            logger.info("开始读取Excel文件...")
            df = pd.read_excel(excel_file2, sheet_name=sheet_name)
            logger.info(f"工作表 {sheet_name} 数据形状: {df.shape}")

            if df.empty:
                results2[sheet_name] = "跳过（空表）"
                logger.info(f"工作表 {sheet_name} 为空，跳过处理")
                continue

            logger.info("开始清理列名...")
            # 使用原始的第一行作为列名，但需要清理以支持中文
            df.columns = [sanitize_name(col, is_column=True) for col in df.columns]
            logger.info(f"清理后的列名: {df.columns.tolist()}")
            
            # 去重列名（避免重复）
            seen = set()
            new_cols = []
            for col in df.columns:
                suffix = 1
                orig = col
                while col in seen:
                    col = f"{orig}_{suffix}"
                    suffix += 1
                seen.add(col)
                new_cols.append(col)
            df.columns = new_cols

            # 强制将第二个文件的第一个工作表导入为 'sheet1' 表
            table_name = "sheet1" if idx == 0 else f"sheet1_{idx}"
            logger.info(f"表名: {table_name}")
            
            # 获取数据库元数据
            metadata = MetaData()
            inspector = inspect(engine)

            # 删除已存在的同名表
            if inspector.has_table(table_name):
                logger.info(f"删除已存在的表: {table_name}")
                old_table = Table(table_name, metadata, autoload_with=engine)
                old_table.drop(engine)
                logger.info(f"成功删除表: {table_name}")

            logger.info("开始推断列类型并构建表结构...")
            # 推断列类型并构建表结构
            columns = []
            type_mapping = {}
            
            for col in df.columns:
                # 推断合适的SQL类型而不是全部使用TEXT
                sql_type = infer_sql_type(df[col])
                columns.append(Column(col, sql_type))
                # 构建 SQLAlchemy 类型映射（用于 to_sql）
                type_mapping[col] = sql_type

            # 创建新表
            table = Table(
                table_name, metadata, *columns,
                mysql_charset='utf8mb4',
                extend_existing=True,
                mysql_row_format='DYNAMIC',
                mysql_engine='InnoDB'
            )
            logger.info(f"创建表 {table_name}")
            metadata.create_all(engine)
            logger.info(f"成功创建表: {table_name}")

            logger.info("开始数据预处理...")
            # 数据预处理 - 确保所有数据都符合对应的SQL类型
            df = preprocess_data(df, type_mapping)
            logger.info("数据预处理完成")

            logger.info("开始插入数据...")
            # 插入数据
            logger.info(f"准备插入 {len(df)} 行数据到表 {table_name}")
            
            # 按表拆分数据 - 由于我们已经简化了处理逻辑，不再需要拆分
            logger.info(f"插入 {len(df)} 行数据到表 {table_name}")
            
            # 获取当前表的类型映射
            current_type_mapping = {}
            for col_name in df.columns:
                if col_name in type_mapping:
                    current_type_mapping[col_name] = type_mapping[col_name]
            
            # 分批插入数据，避免一次性插入大量数据
            batch_size = 1000
            total_rows = len(df)
            
            for start_idx in range(0, total_rows, batch_size):
                end_idx = min(start_idx + batch_size, total_rows)
                batch_df = df.iloc[start_idx:end_idx]
                
                logger.info(f"正在插入批次 {start_idx//batch_size + 1}/{(total_rows-1)//batch_size + 1} 到表 {table_name}")
                
                append_dataframe_to_table(table_name, batch_df)
            
            results2[sheet_name] = f"成功导入 {len(df)} 行数据到表 `{table_name}`"
            logger.info(f"成功插入数据到表: {table_name}")
            
            # 只处理第一个工作表，其他工作表跳过
            if idx == 0:
                break

        # 检查是否成功导入了两个表
        if len(results1) == 0 or len(results2) == 0:
            logger.error("至少有一个文件没有成功导入任何表")
            raise HTTPException(status_code=500, detail="至少有一个文件没有成功导入任何表")

        logger.info("开始执行SQL语句")
        
        # 检查表中是否包含必要的列
        with engine.connect() as conn:
            # 获取固定表名的列信息
            main_table = "sheet"  # 固定主表名
            join_table = "sheet1" # 固定关联表名
            
            logger.info(f"使用固定表名: 主表={main_table}, 关联表={join_table}")
            
            # 获取主表的列信息
            try:
                main_table_columns = conn.execute(text(f"SHOW COLUMNS FROM `{main_table}`")).fetchall()
                main_columns = [col[0] for col in main_table_columns]
                
                # 获取关联表的列信息
                join_table_columns = conn.execute(text(f"SHOW COLUMNS FROM `{join_table}`")).fetchall()
                join_columns = [col[0] for col in join_table_columns]
                
                logger.info(f"主表 {main_table} 的列: {main_columns}")
                logger.info(f"关联表 {join_table} 的列: {join_columns}")
                
                # 检查必需的列是否存在
                # 根据错误信息，主表(sheet)应包含'零件号'和'ESO_Actual_Date'，关联表(sheet1)应包含'零件号'和'归档日期'
                required_main_cols = ['零件号', 'ESO_Plan_Date', 'ESO_Actual_Date']
                required_join_cols = ['零件号', '归档日期']
                
                missing_main_cols = [col for col in required_main_cols if col not in main_columns]
                missing_join_cols = [col for col in required_join_cols if col not in join_columns]
                
                if missing_main_cols or missing_join_cols:
                    error_msg = f"缺少必要的列: "
                    if missing_main_cols:
                        error_msg += f"主表({main_table})缺少: {missing_main_cols}; "
                    if missing_join_cols:
                        error_msg += f"关联表({join_table})缺少: {missing_join_cols}"
                    logger.error(error_msg)
                    
                    update_result = f"更新失败: {error_msg}"
                    select_result = {
                        "row_count": 0,
                        "total_empty_count": 0,
                        "data": [],
                        "error": f"缺少必要的列: {error_msg}"
                    }
                else:
                    updated_row_count = 0
                    # SQL 1: 更新非空归档日期到ESO Actual Date - 使用固定SQL
                    update_sql = """
                    UPDATE sheet
                    JOIN (
                        SELECT 零件号, MAX(归档日期) AS 归档日期
                        FROM sheet1
                        WHERE 归档日期 IS NOT NULL AND TRIM(归档日期) != ''
                        GROUP BY 零件号
                    ) archived_sheet1 ON sheet.零件号 = archived_sheet1.零件号
                    SET sheet.ESO_Actual_Date = archived_sheet1.归档日期
                    WHERE archived_sheet1.归档日期 IS NOT NULL
                      AND TRIM(archived_sheet1.归档日期) != ''
                      AND (sheet.操作类型 IS NULL OR UPPER(TRIM(sheet.操作类型)) != 'D');
                    """
                    
                    try:
                        result = conn.execute(text(update_sql))
                        conn.commit()
                        updated_row_count = result.rowcount
                        logger.info(f"更新SQL执行成功，受影响行数: {result.rowcount}")
                        update_result = f"成功更新 {result.rowcount} 条记录"
                    except Exception as e:
                        logger.error(f"更新SQL执行失败: {str(e)}")
                        update_result = f"更新失败: {str(e)}"
                    
                    try:
                        select_result = build_unfinished_select_result(conn, resolved_target_date, updated_row_count)
                    except Exception as e:
                        logger.error(f"查询SQL执行失败: {str(e)}")
                        select_result = {
                            "row_count": 0,
                            "total_empty_count": 0,
                            "data": [],
                            "error": f"查询失败: {str(e)}"
                        }
            except Exception as e:
                logger.error(f"检查表结构时出错: {str(e)}")
                update_result = f"检查表结构失败: {str(e)}"
                select_result = {
                    "row_count": 0,
                    "total_empty_count": 0,
                    "data": [],
                    "error": f"检查表结构失败: {str(e)}"
                }

        logger.info(f"两个文件处理完成，SQL执行完成")

        return {
            "file1": file1.filename,
            "file2": file2.filename,
            "target_date": resolved_target_date.isoformat(),
            "sheets": len(sheet_names1) + len(sheet_names2),
            "results": {**results1, **results2},
            "modified_workbook": modified_workbook,
            "sql_results": {
                "update_result": update_result,
                "select_result": select_result
            }
        }

    except Exception as e:
        logger.error(f"处理失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"处理失败: {str(e)}")

@app.post("/unfinished-list/", summary="按统计日期重新生成未完成清单")
async def regenerate_unfinished_list(target_date: str = Body(None, embed=True)):
    """
    基于当前服务内存中的最新 ESO 处理结果，按统计日期重新计算未完成清单。
    MySQL 不再是重新生成清单的前置依赖。
    """
    resolved_target_date = resolve_target_date(target_date)
    logger.info(f"重新生成未完成清单，统计日期: {resolved_target_date.isoformat()}")

    try:
        context = LAST_ANALYSIS.get("eso")
        if not context or context.get("df") is None:
            raise HTTPException(status_code=400, detail="请先上传并处理 ESO Excel 文件，再重新生成未完成清单")

        previous_summary = context.get("select_result", {}).get("summary", {})
        select_result = build_eso_select_result_from_dataframe(
            context["df"],
            resolved_target_date,
            matched_completed_count=previous_summary.get("matched_completed_count"),
            archive_df=context.get("archive_df"),
        )
        context["select_result"] = select_result

        return {
            "target_date": resolved_target_date.isoformat(),
            "select_result": select_result,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"重新生成未完成清单失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"重新生成未完成清单失败: {str(e)}")

@app.post("/upload-drawing-excel/", summary="上传图纸相关Excel并生成未完成清单")
async def upload_drawing_excel(
    club_file: UploadFile = File(...),
    published_file: UploadFile | None = File(None),
    target_date: str = Form(None),
):
    """
    图纸场景独立入口。

    - 第一份文件：零件俱乐部或包含数模/图纸计划与实际日期的主清单
    - 第二份文件：可选，图纸发布/归档清单，用于补充图纸实际发布日期
    - 当前按会议规则生成初版：操作类型 D 不进入未完成清单；数模未完成时数模和图纸都入未完成；数模完成后一个月内图纸未发布则入未完成
    """
    if not club_file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .xls 文件 (club_file)")
    if published_file and not published_file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .xls 文件 (published_file)")

    try:
        resolved_target_date = resolve_target_date(target_date)
        club_contents = await club_file.read()
        published_contents = await published_file.read() if published_file else None
        result, modified_workbook, persistence = process_drawing_files(
            club_contents,
            published_contents,
            club_file.filename,
            published_file.filename if published_file else None,
            resolved_target_date,
        )
        return {
            "club_file": club_file.filename,
            "published_file": published_file.filename if published_file else None,
            "target_date": resolved_target_date.isoformat(),
            "select_result": result,
            "modified_workbook": modified_workbook,
            "persistence": persistence,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"图纸未完成清单生成失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"图纸未完成清单生成失败: {str(e)}")

@app.post("/drawing-unfinished-list/", summary="按统计日期重新生成图纸未完成清单")
async def regenerate_drawing_unfinished_list(target_date: str = Body(None, embed=True)):
    """
    基于当前服务内存中的最新图纸文件，按统计日期重新计算图纸未完成清单。
    不重新上传 Excel，不重新生成回填文件。
    """
    resolved_target_date = resolve_target_date(target_date)
    logger.info(f"重新生成图纸未完成清单，统计日期: {resolved_target_date.isoformat()}")

    try:
        context = LAST_ANALYSIS.get("drawing")
        if not context or context.get("club_bytes") is None:
            raise HTTPException(status_code=400, detail="请先上传并处理图纸 Excel 文件，再重新生成未完成清单")

        select_result = build_drawing_unfinished_result(
            context["club_bytes"],
            context.get("published_bytes"),
            resolved_target_date,
        )
        previous_summary = context.get("select_result", {}).get("summary", {})
        if previous_summary.get("matched_completed_count") is not None:
            select_result["summary"]["matched_completed_count"] = previous_summary.get("matched_completed_count")
        if previous_summary.get("delete_skipped_count") is not None:
            select_result["summary"]["delete_skipped_count"] = previous_summary.get("delete_skipped_count")
        context["select_result"] = select_result

        return {
            "target_date": resolved_target_date.isoformat(),
            "select_result": select_result,
            "modified_workbook": context.get("modified_workbook"),
            "persistence": context.get("persistence"),
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"重新生成图纸未完成清单失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"重新生成图纸未完成清单失败: {str(e)}")

def load_latest_standard_result_from_mysql(scene: str):
    try:
        with engine.connect() as conn:
            batch = conn.execute(text("""
            SELECT batch_id, summary_json
            FROM analysis_batches
            WHERE scene = :scene
            ORDER BY created_at DESC
            LIMIT 1
            """), {"scene": scene}).mappings().first()
            if not batch:
                return None

            item_rows = conn.execute(text("""
            SELECT payload_json
            FROM analysis_unfinished_items
            WHERE scene = :scene AND batch_id = :batch_id
            ORDER BY row_index ASC
            """), {"scene": scene, "batch_id": batch["batch_id"]}).mappings().all()
            group_rows = conn.execute(text("""
            SELECT group_name, count_value
            FROM analysis_group_stats
            WHERE scene = :scene AND batch_id = :batch_id
            ORDER BY count_value DESC
            """), {"scene": scene, "batch_id": batch["batch_id"]}).mappings().all()

            data = [json.loads(row["payload_json"]) for row in item_rows]
            summary = json.loads(batch["summary_json"] or "{}")
            group_stats = [{"功能组": row["group_name"], "count": int(row["count_value"])} for row in group_rows]
            return {
                "data": data,
                "row_count": len(data),
                "total_empty_count": len(data),
                "summary": summary,
                "group_stats": group_stats,
            }
    except Exception as e:
        logger.info("从MySQL读取最新标准结果失败: %s", str(e))
        return None

def get_latest_select_result(scene: str):
    context = LAST_ANALYSIS.get(scene)
    if context and context.get("select_result"):
        return context["select_result"], "memory"

    result = load_latest_standard_result_from_mysql(scene)
    if result:
        return result, "mysql"
    return None, None

def infer_scene(question: str, requested_scene: str | None = None):
    if requested_scene in {"eso", "drawing"}:
        return requested_scene
    lower_question = question.lower()
    if "图纸" in question or "数模" in question or "drawing" in lower_question:
        return "drawing"
    return "eso"

def make_table_data(records: List[Dict[str, Any]]):
    columns = []
    if records:
        for key in records[0].keys():
            columns.append({
                "prop": key,
                "label": key,
                "width": min(max(len(str(key)) * 10, 100), 220),
            })
    return {
        "records": records,
        "columns": columns,
        "total": len(records),
    }

def answer_smart_question(question: str, scene: str | None = None):
    resolved_scene = infer_scene(question, scene)
    select_result, source = get_latest_select_result(resolved_scene)
    if not select_result:
        return {
            "answer": "当前还没有可查询的数据。请先在对应页面上传并生成未完成清单；如果需要历史查询，请确认 MySQL 已启动并已保存过结果。",
            "scene": resolved_scene,
            "source": None,
            "table_data": make_table_data([]),
        }

    rows = select_result.get("data", [])
    summary = select_result.get("summary", {})
    question_text = question.strip()
    lower_question = question_text.lower()
    wants_engineer = "工程师" in question_text or "engineer" in lower_question
    wants_group = "功能组" in question_text or "部门" in question_text or "group" in lower_question
    wants_type = "类型" in question_text or "数模" in question_text or "图纸" in question_text
    wants_part_no = "零件号" in question_text or "part" in lower_question
    wants_count = any(token in question_text for token in ["统计", "数量", "多少", "几个", "汇总"])
    wants_detail = any(token in question_text for token in ["明细", "清单", "告诉我", "有哪些", "列出", "全部"])

    if wants_engineer and wants_count:
        counter: Dict[str, int] = {}
        for row in rows:
            key = str(row.get("工程师") or "未知工程师")
            counter[key] = counter.get(key, 0) + 1
        records = [{"工程师": key, "未完成数量": value} for key, value in sorted(counter.items(), key=lambda item: item[1], reverse=True)]
        answer = f"{'ESO' if resolved_scene == 'eso' else '图纸'}未完成按工程师统计共 {len(records)} 位工程师，合计 {len(rows)} 项。"
        return {"answer": answer, "scene": resolved_scene, "source": source, "table_data": make_table_data(records)}

    if wants_group and wants_count:
        records = [
            {"功能组": item.get("功能组") or item.get("部门") or "未知功能组", "未完成数量": item.get("count", 0)}
            for item in select_result.get("group_stats", [])
        ]
        if not records:
            counter: Dict[str, int] = {}
            for row in rows:
                key = str(row.get("功能组") or row.get("部门") or "未知功能组")
                counter[key] = counter.get(key, 0) + 1
            records = [{"功能组": key, "未完成数量": value} for key, value in sorted(counter.items(), key=lambda item: item[1], reverse=True)]
        answer = f"{'ESO' if resolved_scene == 'eso' else '图纸'}未完成按功能组统计合计 {len(rows)} 项。"
        return {"answer": answer, "scene": resolved_scene, "source": source, "table_data": make_table_data(records)}

    if resolved_scene == "drawing" and wants_type and wants_count:
        counter: Dict[str, int] = {}
        for row in rows:
            key = str(row.get("未完成类型") or "未知类型")
            counter[key] = counter.get(key, 0) + 1
        records = [{"未完成类型": key, "数量": value} for key, value in sorted(counter.items(), key=lambda item: item[1], reverse=True)]
        return {"answer": f"图纸未完成按类型统计合计 {len(rows)} 项。", "scene": resolved_scene, "source": source, "table_data": make_table_data(records)}

    if wants_part_no:
        records = [{"零件号": row.get("零件号", "")} for row in rows if row.get("零件号")]
        answer = f"{'ESO' if resolved_scene == 'eso' else '图纸'}当前未完成零件号共 {len(records)} 个。"
        return {"answer": answer, "scene": resolved_scene, "source": source, "table_data": make_table_data(records)}

    if wants_count and not wants_detail:
        records = [{
            "场景": "ESO" if resolved_scene == "eso" else "图纸",
            "统计日期": summary.get("target_date", ""),
            "未完成数量": summary.get("unfinished_count", len(rows)),
            "统计口径": summary.get("formula", ""),
        }]
        return {"answer": f"当前未完成数量为 {summary.get('unfinished_count', len(rows))} 项。", "scene": resolved_scene, "source": source, "table_data": make_table_data(records)}

    default_columns = ["零件号", "功能组", "工程师", "ESO_Plan_Date", "ESO_Actual_Date", "ESO状态"] if resolved_scene == "eso" else ["零件号", "功能组", "工程师", "未完成类型", "数模状态", "图纸状态", "图纸要求完成日期"]
    records = []
    for row in rows:
        records.append({col: row.get(col, "") for col in default_columns if col in row})
    answer = f"已查询到{'ESO' if resolved_scene == 'eso' else '图纸'}未完成清单 {len(records)} 条。"
    return {"answer": answer, "scene": resolved_scene, "source": source, "table_data": make_table_data(records)}

@app.post("/smart-query/", summary="受控智能问答")
async def smart_query(
    question: str = Body(..., embed=True),
    scene: str = Body(None, embed=True),
):
    if not question or not question.strip():
        raise HTTPException(status_code=400, detail="问题不能为空")
    try:
        return answer_smart_question(question, scene)
    except Exception as e:
        logger.error(f"智能问答失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"智能问答失败: {str(e)}")

@app.get("/download/{filename}", summary="下载生成的Excel文件")
async def download_generated_file(filename: str):
    safe_name = Path(filename).name
    file_path = OUTPUT_DIR / safe_name
    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(status_code=404, detail="文件不存在或已被清理")

    return FileResponse(
        path=file_path,
        filename=safe_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@app.post("/execute-sql/")
async def execute_sql(sql_query: str = Body(..., embed=True)):
    """
    执行SQL语句
    
    - **sql_query**: 要执行的SQL语句
    """
    logger.info(f"收到SQL执行请求: {sql_query}")
    
    # 检查SQL语句是否为空
    if not sql_query or not sql_query.strip():
        logger.warning("SQL语句为空")
        raise HTTPException(status_code=400, detail="SQL语句不能为空")
    
    # 检查SQL是否包含危险操作
    dangerous_keywords = ["drop", "delete", "truncate", "alter", "create"]
    sql_lower = sql_query.strip().lower()
    for keyword in dangerous_keywords:
        if keyword in sql_lower and keyword == sql_lower.split()[0]:  # 检查是否是语句开头的关键词
            logger.warning(f"检测到潜在危险SQL操作: {sql_query}")
            raise HTTPException(status_code=400, detail=f"不允许执行{keyword.upper()}操作")
    
    try:
        logger.info(f"执行SQL查询: {sql_query}")
        # 执行SQL查询
        with engine.connect() as conn:
            result = conn.execute(text(sql_query))
            
            # 判断是否是查询语句
            if sql_query.strip().upper().startswith("SELECT"):
                # 获取列名
                columns = result.keys()
                # 获取所有行数据
                rows = [dict(zip(columns, row)) for row in result.fetchall()]
                
                logger.info(f"查询成功，返回 {len(rows)} 行数据")
                return {
                    "success": True,
                    "row_count": len(rows),
                    "data": rows,
                    "sql_query": sql_query
                }
            else:
                # 对于非查询语句，提交事务并返回受影响的行数
                conn.commit()
                logger.info(f"执行成功，受影响行数: {result.rowcount}")
                return {
                    "success": True,
                    "row_count": result.rowcount,
                    "sql_query": sql_query
                }
    except Exception as e:
        logger.error(f"执行SQL失败: {str(e)}", exc_info=True)
        logger.error(f"SQL语句: {sql_query}")
        logger.error(f"错误详情: {traceback.format_exc()}")
        
        # 返回更具体的错误信息
        error_detail = str(e).lower()
        if "unknown table" in error_detail or "table" in error_detail and "doesn't exist" in error_detail:
            return {
                "success": False,
                "error": "表不存在，请检查表名是否正确",
                "sql_query": sql_query
            }
        elif "unknown column" in error_detail:
            return {
                "success": False,
                "error": f"字段不存在: {str(e)}",
                "sql_query": sql_query
            }
        elif "syntax error" in error_detail or "you have an error in your sql syntax" in error_detail:
            return {
                "success": False,
                "error": f"SQL语法错误: {str(e)}",
                "sql_query": sql_query
            }
        else:
            return {
                "success": False,
                "error": f"执行SQL失败: {str(e)}",
                "sql_query": sql_query
            }

@app.get("/", summary="根路径")
async def root():
    logger.info("访问根路径")
    return {"message": "Excel to MySQL 服务正在运行", "status": "ok"}

@app.get("/health", summary="健康检查")
async def health_check():
    logger.info("健康检查请求")
    try:
        # 测试数据库连接
        with engine.connect() as conn:
            result = conn.execute(text("SELECT 1"))
            logger.info("数据库连接测试成功")
        return {"status": "healthy", "database": "connected"}
    except Exception as e:
        logger.warning(f"数据库连接测试失败，核心Excel处理仍可用: {str(e)}")
        return {"status": "healthy", "database": "optional_unavailable", "database_error": str(e)}

@app.on_event("startup")
async def startup_event():
    logger.info("服务启动中...")
    try:
        # 测试数据库连接
        with engine.connect() as conn:
            result = conn.execute(text("SELECT 1"))
        logger.info("数据库连接初始化成功")
    except Exception as e:
        logger.warning(f"数据库连接初始化失败，MySQL持久化和跨重启问答不可用: {str(e)}")
        
@app.on_event("shutdown")
async def shutdown_event():
    logger.info("服务关闭中...")
    try:
        engine.dispose()
        logger.info("数据库连接已关闭")
    except Exception as e:
        logger.error(f"关闭数据库连接时出错: {str(e)}")


if __name__ == "__main__":
    logger.info("启动Excel to MySQL服务...")
   
    try:
        # 启动服务
        uvicorn.run("main:app", host="0.0.0.0", port=8020, reload=False)
    except Exception as e:
        logger.error(f"启动服务时出错: {str(e)}", exc_info=True)
